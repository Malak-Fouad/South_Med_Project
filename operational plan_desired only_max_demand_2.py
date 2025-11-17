import math
import pandas as pd
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook
import warnings


warnings.filterwarnings('ignore')

# run it second in south med
class RouteAnalyzer:
    def __init__(self, line_name, route_data, dwell_time=3):
        self.line_name = line_name
        self.route_data = route_data
        self.dwell_time = dwell_time
    
    def extract_stops_from_route(self, route_string):
        """Extract stop numbers from route string like '747 → 3972 → 3970 → 3968 → 748'"""
        if pd.isna(route_string):
            return []
        stops = [stop.strip() for stop in route_string.split('→')]
        return [stop for stop in stops if stop]
    
    def convert_runtime_to_minutes(self, runtime_str):
        """Convert LINKRUNTIME from seconds to minutes"""
        try:
            # Remove 's' if present and convert to float
            if isinstance(runtime_str, str):
                runtime_str = runtime_str.replace('s', '').strip()
            runtime_seconds = float(runtime_str)
            # Convert to minutes
            return runtime_seconds / 60
        except (ValueError, TypeError):
            return 0
    
    def calculate_cycle_time(self, route_data):
        """Calculate cycle time based on LINKRUNTIME (converted to minutes) and number of stops"""
        total_runtime = 0
        total_stops = 0
        
        for route in route_data:
            runtime_minutes = self.convert_runtime_to_minutes(route['LINKRUNTIME'])
            total_runtime += runtime_minutes
            
            stops = self.extract_stops_from_route(route['StopsArray'])
            total_stops += len(stops)
        
        cycle_time = total_runtime + (total_stops * self.dwell_time)
        return cycle_time
    
    def get_route_demands(self, route_data):
        """Get individual route demands and max demand"""
        route_demands = {}
        desired_demand = 0
        
        for i, route in enumerate(route_data, 1):
            route_name = route['LINEROUTENAME']
            demand = route['VOL_AP_MAX']
            route_demands[f'Route_{i}_Demand'] = demand
            route_demands[f'Route_{i}_Name'] = route_name
            desired_demand = max(desired_demand, demand)
        
        route_demands['Desired_Demand'] = desired_demand
        return route_demands
    
    def analyze_system_with_headway(self, desired_demand, cycle_time, headway, bus_capacity):
        """Analyze the complete system based on desired demand"""
        if desired_demand > 0 and cycle_time > 0:
            groups_per_hour = math.ceil(60 / headway)
            required_capacity_per_group = desired_demand / groups_per_hour
            buses_per_group = math.ceil(required_capacity_per_group / bus_capacity)
            total_trips = math.ceil(desired_demand / bus_capacity)
            actual_capacity_per_group = buses_per_group * bus_capacity
            total_capacity_per_hour = actual_capacity_per_group * groups_per_hour
            groups_in_service = cycle_time / headway
            unique_groups = math.ceil(groups_in_service)
            fleet_size_performing_Headway_for_1_Hour = min(groups_per_hour, unique_groups) * buses_per_group  
            hub_area_for_1_hour = fleet_size_performing_Headway_for_1_Hour * 70
            empty_seats = total_capacity_per_hour - desired_demand
        else: 
            buses_per_group = 0
            fleet_size_performing_Headway_for_1_Hour = 0
            total_capacity_per_hour = 0
            empty_seats = 0
            groups_per_hour = 0
            unique_groups = 0
            total_trips = 0
            hub_area_for_1_hour = 0
        
        return {
            'buses_per_group': buses_per_group,
            'total_trips' : total_trips,
            'fleet_size_performing_Headway_for_1_Hour': fleet_size_performing_Headway_for_1_Hour,
            'hub_area_for_1_hour': hub_area_for_1_hour,
            'total_capacity_per_hour': total_capacity_per_hour,
            'empty_seats': empty_seats,
            'groups_per_hour': groups_per_hour,
            'unique_groups': unique_groups
        }
    
    def format_route_display(self, route_string):
        """Format route for display"""
        return route_string.replace('→', ' → ')

class DarkExcelStopProcessor:
    def __init__(self, root):
        self.root = root
        self.root.title("South Med Operational Plan Generator - Dark Theme")
        self.root.geometry("950x700")
        self.root.configure(bg='#2b2b2b')
        
        self.file_path = None
        self.data = None
        self.processed_lines = {}
        self.output_dir = None
        
        self.configure_dark_theme()
        self.create_widgets()
    
    def configure_dark_theme(self):
        style = ttk.Style()
        
        bg_color = '#2b2b2b'
        dark_grey = '#3c3f41'
        medium_grey = '#555555'
        light_grey = '#bbbbbb'
        accent_color = '#4ec9b0'
        
        style.theme_use('default')
        
        style.configure('TFrame', background=bg_color)
        style.configure('TLabel', background=bg_color, foreground=light_grey, font=('Arial', 10))
        style.configure('TButton', background=dark_grey, foreground=light_grey, 
                       font=('Arial', 10, 'bold'), borderwidth=1, focusthickness=3, focuscolor='none')
        style.configure('TLabelframe', background=bg_color, foreground=accent_color, font=('Arial', 11, 'bold'))
        style.configure('TLabelframe.Label', background=bg_color, foreground=accent_color)
        
        style.configure('Treeview', 
                       background=dark_grey,
                       foreground=light_grey,
                       fieldbackground=dark_grey,
                       font=('Arial', 9))
        
        style.configure('Treeview.Heading', 
                       background=medium_grey,
                       foreground=light_grey,
                       font=('Arial', 10, 'bold'))
        
        style.map('Treeview', background=[('selected', '#0d3d56')])
        style.map('TButton', background=[('active', '#555555')])
    
    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="15")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        title_label = tk.Label(main_frame, text="South Med Operational Plan Generator", 
                              font=('Arial', 18, 'bold'), 
                              bg='#2b2b2b', fg='#4ec9b0')
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        file_frame = ttk.LabelFrame(main_frame, text="📁 SELECT EXCEL FILE", padding="12")
        file_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 15))
        
        self.file_label = ttk.Label(file_frame, text="No file selected", font=('Arial', 9))
        self.file_label.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 15))
        
        browse_btn = ttk.Button(file_frame, text="Browse Excel File", 
                               command=self.browse_file)
        browse_btn.grid(row=0, column=1, padx=(0, 10))
        
        load_btn = ttk.Button(file_frame, text="📊 Load Data", 
                             command=self.load_data)
        load_btn.grid(row=0, column=2)
        
        output_frame = ttk.LabelFrame(main_frame, text="📂 SELECT OUTPUT FOLDER", padding="12")
        output_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 15))
        
        self.output_label = ttk.Label(output_frame, text="No output folder selected", font=('Arial', 9))
        self.output_label.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 15))
        
        output_browse_btn = ttk.Button(output_frame, text="Choose Output Folder", 
                                      command=self.choose_output_folder)
        output_browse_btn.grid(row=0, column=1)
        
        config_frame = ttk.LabelFrame(main_frame, text="⚙️ CONFIGURATION", padding="12")
        config_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 15))
        
        ttk.Label(config_frame, text="Bus Capacities:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        self.bus_capacities_var = tk.StringVar(value="25, 50")
        bus_capacities_entry = ttk.Entry(config_frame, textvariable=self.bus_capacities_var, width=20)
        bus_capacities_entry.grid(row=0, column=1, sticky=tk.W, padx=(0, 20))
        
        ttk.Label(config_frame, text="Headways (min):").grid(row=0, column=2, sticky=tk.W, padx=(0, 10))
        self.headways_var = tk.StringVar(value="10, 15, 20, 25, 30")
        headways_entry = ttk.Entry(config_frame, textvariable=self.headways_var, width=20)
        headways_entry.grid(row=0, column=3, sticky=tk.W, padx=(0, 20))
        
        ttk.Label(config_frame, text="Dwell Time (min):").grid(row=0, column=4, sticky=tk.W, padx=(0, 10))
        self.dwell_time_var = tk.StringVar(value="3")
        dwell_time_entry = ttk.Entry(config_frame, textvariable=self.dwell_time_var, width=10)
        dwell_time_entry.grid(row=0, column=5, sticky=tk.W)
        
        process_btn = ttk.Button(main_frame, text="🚀 GENERATE OPERATIONAL PLANS", 
                                command=self.generate_operational_plans)
        process_btn.grid(row=4, column=0, columnspan=3, pady=15, ipadx=20, ipady=5)
        
        results_frame = ttk.LabelFrame(main_frame, text="📊 PROCESSED LINES", padding="12")
        results_frame.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 15))
        
        self.tree = ttk.Treeview(results_frame, 
                                columns=("$LINEROUTEITEM:LINENAME", "Routes", "HubName", "Route1_Demand", "Route2_Demand", "MaxDemand", "CycleTime"), 
                                show="headings", height=12)
        self.tree.heading("$LINEROUTEITEM:LINENAME", text="LINE NAME")
        self.tree.heading("Routes", text="ROUTES")
        self.tree.heading("HubName", text="HUB NAME")
        self.tree.heading("Route1_Demand", text="ROUTE 1 DEMAND")
        self.tree.heading("Route2_Demand", text="ROUTE 2 DEMAND")
        self.tree.heading("MaxDemand", text="MAX DEMAND")
        self.tree.heading("CycleTime", text="CYCLE TIME (min)")
        
        self.tree.column("$LINEROUTEITEM:LINENAME", width=150)
        self.tree.column("Routes", width=200)
        self.tree.column("HubName", width=100)
        self.tree.column("Route1_Demand", width=120)
        self.tree.column("Route2_Demand", width=120)
        self.tree.column("MaxDemand", width=100)
        self.tree.column("CycleTime", width=120)
        
        scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.grid(row=6, column=0, columnspan=3, pady=10)
        
        export_btn = ttk.Button(buttons_frame, text="💾 EXPORT ALL PLANS", 
                               command=self.export_all_plans)
        export_btn.grid(row=0, column=0, padx=(0, 10))
        
        clear_btn = ttk.Button(buttons_frame, text="🗑️ CLEAR RESULTS", 
                              command=self.clear_results)
        clear_btn.grid(row=0, column=1, padx=(0, 10))
        
        self.status_var = tk.StringVar()
        self.status_var.set("Ready to process Excel file")
        status_bar = tk.Label(main_frame, textvariable=self.status_var, 
                             font=('Arial', 8), bg='#2b2b2b', fg='#888888')
        status_bar.grid(row=7, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(10, 0))
        
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(5, weight=1)
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)
        file_frame.columnconfigure(0, weight=1)
        output_frame.columnconfigure(0, weight=1)
        config_frame.columnconfigure(0, weight=0)
        config_frame.columnconfigure(1, weight=0)
        config_frame.columnconfigure(2, weight=0)
        config_frame.columnconfigure(3, weight=0)
        config_frame.columnconfigure(4, weight=0)
        config_frame.columnconfigure(5, weight=1)
    
    def browse_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if file_path:
            self.file_path = file_path
            filename = os.path.basename(file_path)
            self.file_label.config(text=filename)
            self.status_var.set(f"File selected: {filename}")
    
    def choose_output_folder(self):
        folder_path = filedialog.askdirectory(
            title="Select Output Folder for Operational Plans"
        )
        
        if folder_path:
            self.output_dir = folder_path
            self.output_label.config(text=folder_path)
            self.status_var.set(f"Output folder selected: {folder_path}")
    
    def load_data(self):
        if not self.file_path:
            messagebox.showerror("Error", "Please select an Excel file first.")
            return
        
        try:
            self.data = pd.read_excel(self.file_path)
            
            required_columns = ['$LINEROUTEITEM:LINENAME', 'LINEROUTENAME', 'StopsArray', 'HubName', 'LINKRUNTIME', 'MAX:LINEROUTEITEMS\VOL(AP)']
            missing_columns = [col for col in required_columns if col not in self.data.columns]
            
            if missing_columns:
                messagebox.showerror("Error", f"Missing required columns: {', '.join(missing_columns)}")
                return
            
            self.processed_lines = {}
            
            for line_name in self.data['$LINEROUTEITEM:LINENAME'].unique():
                line_data = self.data[self.data['$LINEROUTEITEM:LINENAME'] == line_name]
                routes = []
                
                for _, row in line_data.iterrows():
                    route_info = {
                        'LINEROUTENAME': row['LINEROUTENAME'],
                        'StopsArray': row['StopsArray'],
                        'HubName': row['HubName'],
                        'LINKRUNTIME': row['LINKRUNTIME'],
                        'VOL_AP_MAX': row['MAX:LINEROUTEITEMS\VOL(AP)']
                    }
                    routes.append(route_info)
                
                self.processed_lines[line_name] = routes
            
            self.display_processed_lines()
            self.status_var.set(f"Successfully loaded {len(self.processed_lines)} lines")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error loading file: {str(e)}")
            self.status_var.set("Error loading file")
    
    def display_processed_lines(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        for line_name, routes in self.processed_lines.items():
            analyzer = RouteAnalyzer(line_name, routes)
            route_demands = analyzer.get_route_demands(routes)
            cycle_time = analyzer.calculate_cycle_time(routes)
            
            route_display = " | ".join([route['LINEROUTENAME'] for route in routes])
            
            # Get HubName (assuming all routes in a line have the same hub)
            hub_name = routes[0]['HubName'] if routes else 'N/A'
            
            route1_demand = route_demands.get('Route_1_Demand', 'N/A')
            route2_demand = route_demands.get('Route_2_Demand', 'N/A')
            desired_demand = route_demands['Desired_Demand']
            
            self.tree.insert("", "end", values=(
                line_name,
                route_display,
                hub_name,
                f"{route1_demand:,.0f}" if route1_demand != 'N/A' else 'N/A',
                f"{route2_demand:,.0f}" if route2_demand != 'N/A' else 'N/A',
                f"{desired_demand:,.0f}",
                f"{cycle_time:.1f}"
            ))
    
    def generate_operational_plans(self):
        if not self.processed_lines:
            messagebox.showerror("Error", "No data loaded. Please load Excel file first.")
            return
        
        if not self.output_dir:
            messagebox.showerror("Error", "Please select an output folder first.")
            return
        
        try:
            bus_capacities = [int(x.strip()) for x in self.bus_capacities_var.get().split(',')]
            headways = [int(x.strip()) for x in self.headways_var.get().split(',')]
            dwell_time = int(self.dwell_time_var.get())
            
            operational_plans_dir = os.path.join(self.output_dir, "Operational_Plans")
            os.makedirs(operational_plans_dir, exist_ok=True)
            
            generated_files = []
            
            for line_name, routes in self.processed_lines.items():
                analyzer = RouteAnalyzer(line_name, routes, dwell_time)
                
                route_demands = analyzer.get_route_demands(routes)
                cycle_time = analyzer.calculate_cycle_time(routes)
                desired_demand = route_demands['Desired_Demand']
                
                # Get HubName (assuming all routes in a line have the same hub)
                hub_name = routes[0]['HubName'] if routes else 'N/A'
                
                results = []
                
                for bus_capacity in bus_capacities:
                    for headway in headways:
                        system_analysis = analyzer.analyze_system_with_headway(
                            desired_demand, cycle_time, headway, bus_capacity
                        )
                        
                        # Create row_data in the exact order you want
                        row_data = {}
                        
                        # Hub information first
                        row_data['HubName'] = hub_name
                        
                        # Route information
                        for i, route in enumerate(routes, 1):
                            route_name = route['LINEROUTENAME']
                            demand = route['VOL_AP_MAX']
                            row_data[f'Route_{i}_Name'] = route_name
                            row_data[f'Route_{i}_Stops'] = route['StopsArray']
                        
                        # Route demands
                        for i, route in enumerate(routes, 1):
                            row_data[f'Route_{i}_Demand'] = route['VOL_AP_MAX']
                        
                        row_data['Desired_Demand'] = desired_demand
                        row_data['Bus_Capacity'] = bus_capacity
                        row_data['Headway (min)'] = headway
                        row_data['Cycle_Time (min)'] = round(cycle_time, 1)
                        row_data['Buses_per_Group'] = system_analysis['buses_per_group']
                        row_data['Total_Trips'] = system_analysis['total_trips']
                        row_data['Groups_per_Hour'] = system_analysis['groups_per_hour']
                        row_data['Unique_Groups'] = system_analysis['unique_groups']
                        row_data['Fleet_Size'] = system_analysis['fleet_size_performing_Headway_for_1_Hour']
                        
                        row_data['Hub_Area'] = system_analysis['hub_area_for_1_hour']
                       
                        row_data['Capacity_per_Hour'] = round(system_analysis['total_capacity_per_hour'], 1)
                        row_data['Empty_Seats_per_Hour'] = round(system_analysis['empty_seats'], 1)
                        
                        results.append(row_data)
                
                df = pd.DataFrame(results)
                
                # Define the exact column order with HubName first
                final_columns = [
                    'HubName',
                    'Route_1_Name',
                    'Route_1_Stops', 
                    'Route_2_Name',
                    'Route_2_Stops',
                    'Route_1_Demand',
                    'Route_2_Demand',
                    'Desired_Demand',
                    'Bus_Capacity',
                    'Headway (min)',
                    'Cycle_Time (min)',
                    'Buses_per_Group',
                    'Total_Trips',
                    'Groups_per_Hour',
                    'Unique_Groups',
                    'Fleet_Size',
                    'Hub_Area',
                    'Capacity_per_Hour',
                    'Empty_Seats_per_Hour'
                ]
                
                df = df[final_columns]
                
                safe_line_name = "".join(c for c in line_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
                filename = os.path.join(operational_plans_dir, f"Operational_Plan_{safe_line_name}.xlsx")
                
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Operational_Analysis', index=False)
                    
                    ws = writer.sheets['Operational_Analysis']
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)
                    
                    for col in range(1, len(ws[1]) + 1):
                        ws.cell(1, col).fill = header_fill
                        ws.cell(1, col).font = header_font
                        ws.cell(1, col).alignment = Alignment(horizontal='center')
                    
                    for column in ws.columns:
                        max_length = 0
                        col_letter = column[0].column_letter
                        for cell in column:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
                
                generated_files.append(filename)
            
            self.status_var.set(f"Generated {len(generated_files)} operational plans in '{operational_plans_dir}'")
            messagebox.showinfo("Success", 
                            f"Successfully generated {len(generated_files)} operational plan files!\n\n"
                            f"Output folder: {operational_plans_dir}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error generating operational plans: {str(e)}")
            self.status_var.set("Error generating plans")
        
    def export_all_plans(self):
        self.generate_operational_plans()
    
    def clear_results(self):
        self.processed_lines = {}
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.status_var.set("Results cleared")

def main():
    root = tk.Tk()
    app = DarkExcelStopProcessor(root)
    root.mainloop()

if __name__ == "__main__":
    main()